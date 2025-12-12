#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modbus Register Excel Exporter
===============================
Экспортирует данные из SQLite БД в красиво оформленный Excel файл.

Использование:
    python3 export_to_excel.py

Требования:
    pip install openpyxl

Автор: Claude Code
Дата: 2025-12-12
"""

import sqlite3
from pathlib import Path
from typing import List, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Библиотека openpyxl не установлена")
    print("   Установите: pip install openpyxl")
    exit(1)


# Пути к файлам
SCRIPT_DIR = Path(__file__).parent
DB_PATH = SCRIPT_DIR / 'db' / 'modbus_registers.db'
EXCEL_OUTPUT_PATH = SCRIPT_DIR / 'modbus_map.xlsx'

# Цветовая схема
COLOR_HEADER = 'D9D9D9'  # Серый для заголовков
COLOR_BOOL = 'DAEEF3'    # Голубой для BOOL
COLOR_REAL = 'D9EAD3'    # Зелёный для REAL
COLOR_INT = 'FFF2CC'     # Жёлтый для INT/UINT/WORD
COLOR_SECTION = 'F2F2F2' # Светло-серый для разделителей секций


class ExcelExporter:
    """Экспорт данных из БД в Excel с форматированием"""

    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Удалить дефолтный лист

    def export(self):
        """Экспортировать все данные"""
        # Создать листы для каждого типа регистров
        self._export_register_type('holding_registers', 'Holding Registers')
        self._export_register_type('input_registers', 'Input Registers')

        # Создать лист с резервами
        self._export_gaps()

        # Создать сводную статистику
        self._export_statistics()

    def _export_register_type(self, type_name: str, sheet_name: str):
        """Экспортировать регистры определённого типа"""
        ws = self.wb.create_sheet(sheet_name)

        # Заголовки колонок
        headers = ['Адрес', 'Тип данных', 'Переменная PLC', 'Описание', 'Секция']
        ws.append(headers)

        # Форматирование заголовка
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = self._get_border()

        # Получить данные из БД
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT
                r.register_address,
                r.bit_index,
                dt.name AS data_type,
                dt.register_count,
                r.variable_name,
                r.description,
                s.name AS section
            FROM registers r
            JOIN data_types dt ON r.data_type_id = dt.id
            JOIN sections s ON r.section_id = s.id
            JOIN register_types rt ON r.register_type_id = rt.id
            WHERE rt.name = ?
            ORDER BY s.start_register, r.register_address, r.bit_index
        """, (type_name,))

        current_section = None
        row_num = 2

        for row in cursor.fetchall():
            # Добавить разделитель секции
            if current_section != row['section']:
                current_section = row['section']
                self._add_section_separator(ws, row_num, current_section)
                row_num += 1

            # Форматировать адрес
            if row['data_type'] == 'BOOL':
                address = f"{row['register_address']}.{row['bit_index']}"
            elif row['register_count'] == 2:
                address = f"{row['register_address']}-{row['register_address'] + 1}"
            else:
                address = str(row['register_address'])

            # Добавить данные
            data_row = [
                address,
                row['data_type'],
                row['variable_name'],
                row['description'] or '',
                row['section']
            ]
            ws.append(data_row)

            # Форматирование строки
            self._format_data_row(ws, row_num, row['data_type'])
            row_num += 1

        # Автоширина колонок
        self._auto_size_columns(ws)

        # Закрепить верхнюю строку
        ws.freeze_panes = 'A2'

        # Автофильтр
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _export_gaps(self):
        """Экспортировать резервы (gaps)"""
        ws = self.wb.create_sheet('Резервы')

        # Заголовки
        headers = ['Тип регистра', 'Начало', 'Конец', 'Размер', 'Назначение']
        ws.append(headers)

        # Форматирование заголовка
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()

        # Получить данные
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT
                rt.name AS register_type,
                g.start_register,
                g.end_register,
                g.purpose
            FROM register_gaps g
            JOIN register_types rt ON g.register_type_id = rt.id
            ORDER BY rt.id, g.start_register
        """)

        row_num = 2
        for row in cursor.fetchall():
            size = row['end_register'] - row['start_register'] + 1
            data_row = [
                row['register_type'],
                row['start_register'],
                row['end_register'],
                size,
                row['purpose'] or ''
            ]
            ws.append(data_row)
            self._format_data_row(ws, row_num, 'NORMAL')
            row_num += 1

        self._auto_size_columns(ws)
        ws.freeze_panes = 'A2'

    def _export_statistics(self):
        """Экспортировать сводную статистику"""
        ws = self.wb.create_sheet('Статистика', 0)  # Первый лист

        # Заголовок
        ws['A1'] = 'Статистика Modbus регистров'
        ws['A1'].font = Font(bold=True, size=14)

        # Общая статистика
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM registers")
        total_registers = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(DISTINCT section_id) FROM registers")
        total_sections = cursor.fetchone()[0]

        ws['A3'] = 'Общая статистика:'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = f'Всего регистров:'
        ws['B4'] = total_registers
        ws['A5'] = f'Всего секций:'
        ws['B5'] = total_sections

        # Статистика по типам данных
        ws['A7'] = 'Регистры по типам данных:'
        ws['A7'].font = Font(bold=True)

        cursor.execute("""
            SELECT dt.name, COUNT(*) as count
            FROM registers r
            JOIN data_types dt ON r.data_type_id = dt.id
            GROUP BY dt.name
            ORDER BY count DESC
        """)

        row_num = 8
        for row in cursor.fetchall():
            ws[f'A{row_num}'] = f'{row[0]}:'
            ws[f'B{row_num}'] = row[1]
            row_num += 1

        # Статистика по секциям
        ws['A' + str(row_num + 1)] = 'Топ-10 секций по количеству регистров:'
        ws['A' + str(row_num + 1)].font = Font(bold=True)

        cursor.execute("""
            SELECT s.name, COUNT(*) as count
            FROM registers r
            JOIN sections s ON r.section_id = s.id
            GROUP BY s.name
            ORDER BY count DESC
            LIMIT 10
        """)

        row_num += 2
        for row in cursor.fetchall():
            ws[f'A{row_num}'] = row[0]
            ws[f'B{row_num}'] = row[1]
            row_num += 1

        # Автоширина
        self._auto_size_columns(ws)

    def _add_section_separator(self, ws, row_num: int, section_name: str):
        """Добавить разделитель секции"""
        ws.insert_rows(row_num)
        cell = ws.cell(row=row_num, column=1)
        cell.value = f'▼ {section_name}'
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type='solid')

        # Объединить ячейки
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)

    def _format_data_row(self, ws, row_num: int, data_type: str):
        """Форматировать строку данных"""
        # Определить цвет заливки
        if data_type == 'BOOL':
            color = COLOR_BOOL
        elif data_type == 'REAL':
            color = COLOR_REAL
        elif data_type in ('INT', 'UINT', 'WORD', 'DINT'):
            color = COLOR_INT
        else:
            color = 'FFFFFF'

        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        border = self._get_border()

        for col_num in range(1, 6):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    def _get_border(self) -> Border:
        """Получить стиль границ"""
        thin_border = Side(style='thin', color='000000')
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    def _auto_size_columns(self, ws):
        """Автоматически подобрать ширину колонок"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 60)  # Макс 60 символов
            ws.column_dimensions[column_letter].width = adjusted_width

    def save(self, output_path: Path):
        """Сохранить Excel файл"""
        self.wb.save(output_path)

    def close(self):
        """Закрыть соединение с БД"""
        self.conn.close()


def main():
    """Главная функция"""
    print("=" * 60)
    print("Modbus Register Excel Exporter")
    print("=" * 60)

    # Проверка наличия БД
    if not DB_PATH.exists():
        print(f"❌ База данных не найдена: {DB_PATH}")
        print("   Сначала выполните: python3 migrate_from_fc.py")
        return

    # Экспорт данных
    print(f"\n📤 Экспорт из базы данных: {DB_PATH.name}")
    exporter = ExcelExporter(DB_PATH)
    exporter.export()

    # Сохранить файл
    print(f"\n💾 Сохранение в файл: {EXCEL_OUTPUT_PATH.name}")
    exporter.save(EXCEL_OUTPUT_PATH)
    exporter.close()

    print(f"\n✅ Экспорт завершён успешно!")
    print(f"   Файл: {EXCEL_OUTPUT_PATH}")
    print("=" * 60)


if __name__ == '__main__':
    main()
